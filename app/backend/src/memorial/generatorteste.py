import ezdxf
import pandas as pd
import os
import logging
import math
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List, Dict
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from IA_config import *

# ==============================
# CONFIGURAÇÃO E LOGGING
# ==============================
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

# Mapeamento de Layers REAIS (Baseado no teste.dxf)
MAP_LAYER = {
    "ARQ - ALVENARIA ALTA": "parede",
    "ARQUITETÔNICO - ALVENARIA ALTA": "parede",
    "ARQ - ALVENARIA MÉDIA-BAIXA": "parede",
    "ARQ - ESQUADRIAS": "vao",
    "ESQUADRIAS": "vao",
    "ARQ - TEXTOS": "ambiente_nome",
    "ARQUITETÔNICO - TEXTOS": "ambiente_nome",
    "ESTRUTURAL - PILARES": "pilar",
    "ESTRUTURAL - VIGAS": "viga",
    "ESTRUTURAL - LAJES": "laje",
    "HIDROSSANITÁRIO - ÁGUA FRIA": "hidro",
    "HIDROSSANITÁRIO - ESGOTO": "hidro",
    "ELÉTRICA": "eletrica",
}

# ==============================
# MODELOS DE DADOS (MEMORIAL)
# ==============================
@dataclass
class Dimensoes:
    comprimento: Optional[float] = None
    largura: Optional[float] = None
    altura: Optional[float] = 3.0  # Altura padrão (pé-direito)
    espessura: Optional[float] = 0.15

@dataclass
class Vao:
    tipo: Optional[str] = None
    comprimento: Optional[float] = None
    altura: Optional[float] = 2.1
    espessura: Optional[float] = 0.15

@dataclass
class AlvenariaAdicional:
    tipo: Optional[str] = None
    comprimento: Optional[float] = None
    altura: Optional[float] = None
    espessura: Optional[float] = None

@dataclass
class Ambiente:
    nome: str
    dimensoes: Dimensoes = field(default_factory=Dimensoes)
    vao: Vao = field(default_factory=Vao)
    alvenaria_adicional: AlvenariaAdicional = field(default_factory=AlvenariaAdicional)

@dataclass
class ProjetoMemorial:
    nome_projeto: str
    ambientes: List[Ambiente]

# ==============================
# EXTRAÇÃO CAD
# ==============================
class CADExtractor:
    def __init__(self, file_path: str):
        self.classifier = LLMClassifier()
        self.file_path = file_path
        self.doc = self._load_file()

    def _load_file(self):
        try:
            logging.info(f"Carregando arquivo: {self.file_path}")
            return ezdxf.readfile(self.file_path)
        except Exception as e:
            logging.error(f"Erro ao abrir arquivo CAD: {e}")
            return None

    def _calcular_comprimento(self, entity):
        tipo = entity.dxftype()
        try:
            if tipo == 'LINE':
                return (entity.dxf.start - entity.dxf.end).magnitude
            elif tipo == 'LWPOLYLINE':
                pontos = list(entity.get_points())
                if len(pontos) < 2: return 0
                comprimento = 0
                for i in range(len(pontos) - 1):
                    x1, y1 = pontos[i][0], pontos[i][1]
                    x2, y2 = pontos[i + 1][0], pontos[i + 1][1]
                    comprimento += math.hypot(x2 - x1, y2 - y1)
                return comprimento
            elif tipo == 'ARC':
                raio = entity.dxf.radius
                ang = math.radians(entity.dxf.end_angle - entity.dxf.start_angle)
                if ang < 0: ang += 2 * math.pi
                return raio * ang
            elif tipo == 'CIRCLE':
                return 2 * math.pi * entity.dxf.radius
            return 0
        except Exception as e:
            return 0
        
    
    def classificar_elemento(self, entity):
        layer = entity.dxf.layer.upper()
        
        # 1. Heurística (rápida e confiável)
        tipo = MAP_LAYER.get(layer)
        if tipo:
            return tipo

        # 2. IA (fallback)
        nome = getattr(entity.dxf, "name", "")
        texto = ""

        if entity.dxftype() in ["TEXT", "MTEXT"]:
            texto = entity.dxf.text if entity.dxftype() == "TEXT" else entity.text

        tipo_llm, conf = self.classifier.classificar(
            nome=nome,
            layer=layer,
            texto=texto,
            tipo_entidade=entity.dxftype()
        )

        if conf < 0.7:
            return "desconhecido"

        return tipo_llm


    def extrair_dados_reais(self) -> List[Ambiente]:
        if not self.doc: return []
        msp = self.doc.modelspace()
        self.classifier = LLMClassifier()
        
        # 1. Identificar nomes de ambientes (TEXT/MTEXT nas layers mapeadas)
        ambientes_dict: Dict[str, Ambiente] = {}
        for entity in msp.query('TEXT MTEXT'):
            layer = entity.dxf.layer.upper()
            if MAP_LAYER.get(layer) == "ambiente_nome":
                nome = (entity.dxf.text if entity.dxftype() == 'TEXT' else entity.text).strip()
                if nome and len(nome) > 2 and nome not in ambientes_dict:
                    # Limpar formatação MTEXT se houver
                    if '\\P' in nome: nome = nome.split('\\P')[0]
                    ambientes_dict[nome] = Ambiente(nome=nome)

        if not ambientes_dict:
            ambientes_dict["Ambiente Principal"] = Ambiente(nome="Ambiente Principal")

        # 2. Agrupar medidas por tipo de layer (Lógica simplificada para MVP)
        primeiro_ambiente = list(ambientes_dict.values())[0]
        
        for entity in msp:
            layer = entity.dxf.layer.upper()
            tipo_mapeado = self.classificar_elemento(entity)
            

            if tipo_mapeado == "desconhecido":
                continue
            
            if tipo_mapeado == "parede":
                comp = self._calcular_comprimento(entity)
                if not primeiro_ambiente.dimensoes.comprimento: primeiro_ambiente.dimensoes.comprimento = 0
                primeiro_ambiente.dimensoes.comprimento += comp
                logging.info(f"Layer: {layer} → Tipo: {tipo_mapeado}")
                
            elif tipo_mapeado == "vao":
                comp = self._calcular_comprimento(entity)
                if not primeiro_ambiente.vao.comprimento: primeiro_ambiente.vao.comprimento = 0
                primeiro_ambiente.vao.comprimento += comp
                primeiro_ambiente.vao.tipo = "Esquadrias"

        # Limitar para os primeiros 20 ambientes para caber no template
        
        return list(ambientes_dict.values())[:20]
    
    

# ==============================
# GERAÇÃO EXCEL
# ==============================
class LevantamentoCampoMapper:
    START_ROW = 8
    END_ROW = 27

    def __init__(self, worksheet: Worksheet):
        self.ws = worksheet

    def preencher_titulo_projeto(self, nome_projeto: str) -> None:
        self.ws["B1"] = nome_projeto

    def preencher_ambientes(self, ambientes: List[Ambiente]) -> None:
        for index, ambiente in enumerate(ambientes):
            row = self.START_ROW + index
            self.ws[f"B{row}"] = ambiente.nome
            self.ws[f"E{row}"] = round(ambiente.dimensoes.comprimento / 100, 2) if ambiente.dimensoes.comprimento else 0 # Ajuste escala cm->m se necessário
            self.ws[f"G{row}"] = ambiente.dimensoes.altura
            self.ws[f"H{row}"] = ambiente.dimensoes.espessura
            self.ws[f"J{row}"] = ambiente.vao.tipo
            self.ws[f"K{row}"] = round(ambiente.vao.comprimento / 100, 2) if ambiente.vao.comprimento else 0

class MemorialGenerator:
    SHEET_NAME = "Levantamento Campo"

    def __init__(self, template_path: str):
        self.template_path = Path(template_path)

    def generate(self, projeto: ProjetoMemorial, output_path: str) -> Path:
        wb = load_workbook(self.template_path)
        ws = wb[self.SHEET_NAME]
        mapper = LevantamentoCampoMapper(ws)
        mapper.preencher_titulo_projeto(projeto.nome_projeto)
        mapper.preencher_ambientes(projeto.ambientes)
        
        out_p = Path(output_path)
        out_p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_p)
        return out_p

# ==============================
# EXECUÇÃO
# ==============================
def run_integration(dxf_file: str, template_file: str, output_file: str):
    print(f"Processando {os.path.basename(dxf_file)}...")
    
    extractor = CADExtractor(dxf_file)
    ambientes = extractor.extrair_dados_reais()
    
    nome_projeto = os.path.basename(dxf_file).replace(".dxf", "").title()
    projeto = ProjetoMemorial(nome_projeto=nome_projeto, ambientes=ambientes)
    
    generator = MemorialGenerator(template_file)
    arquivo_final = generator.generate(projeto, output_file)
    
    print(f"Sucesso! Memorial gerado em: {arquivo_final}")

if __name__ == "__main__":
    DXF_INPUT = r"C:\Users\faelb\OneDrive\Desktop\EchoCAD\EchoCAD\app\backend\tests\ExtracaoDados\teste.dxf"
    TEMPLATE = r"C:\Users\faelb\OneDrive\Desktop\EchoCAD\EchoCAD\app\backend\src\templates\Memorial de Cálculo - Modelo.xlsx"
    OUTPUT = r"C:\Users\faelb\OneDrive\Desktop\EchoCAD\EchoCAD\app\backend\src\templates\saida\memorial_preenchido.xlsx"
    
    run_integration(DXF_INPUT, TEMPLATE, OUTPUT)
