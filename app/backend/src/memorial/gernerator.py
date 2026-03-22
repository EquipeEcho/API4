from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# modelos de dados que irão para o template
@dataclass
class Dimensoes:
    comprimento: Optional[float] = None
    largura: Optional[float] = None
    altura: Optional[float] = None
    espessura: Optional[float] = None


@dataclass
class Vao:
    tipo: Optional[str] = None
    comprimento: Optional[float] = None
    altura: Optional[float] = None
    espessura: Optional[float] = None


@dataclass
class AlvenariaAdicional:
    tipo: Optional[str] = None
    comprimento: Optional[float] = None
    altura: Optional[float] = None
    espessura: Optional[float] = None


@dataclass
class Ambiente:
    nome: str
    dimensoes: Dimensoes = field(default_factory=Dimensoes)
    vao: Vao = field(default_factory=Vao)
    alvenaria_adicional: AlvenariaAdicional = field(default_factory=AlvenariaAdicional)


@dataclass
class ProjetoMemorial:
    nome_projeto: str
    ambientes: List[Ambiente]


# validação dos dados
class MemorialValidationError(Exception):
    pass


class ProjetoValidator:
    MAX_AMBIENTES = 20

    @classmethod
    def validar(cls, projeto: ProjetoMemorial) -> None:
        if not projeto.nome_projeto or not projeto.nome_projeto.strip():
            raise MemorialValidationError("O nome do projeto é obrigatório.")

        if not projeto.ambientes:
            raise MemorialValidationError("O projeto deve conter ao menos um ambiente.")

        if len(projeto.ambientes) > cls.MAX_AMBIENTES:
            raise MemorialValidationError(
                f"O template suporta no máximo {cls.MAX_AMBIENTES} ambientes na aba 'Levantamento Campo'."
            )

        for i, ambiente in enumerate(projeto.ambientes, start=1):
            if not ambiente.nome or not ambiente.nome.strip():
                raise MemorialValidationError(f"O ambiente #{i} está sem nome.")


# mapper da aba (aqui vai mapear as colunas e linhas do arquivo .xslx)
class LevantamentoCampoMapper:
    START_ROW = 8
    END_ROW = 27

    def __init__(self, worksheet: Worksheet):
        self.ws = worksheet

    def preencher_titulo_projeto(self, nome_projeto: str) -> None:
        self.ws["B1"] = nome_projeto

    def preencher_ambientes(self, ambientes: List[Ambiente]) -> None:
        limite = self.END_ROW - self.START_ROW + 1

        for index, ambiente in enumerate(ambientes[:limite]):
            row = self.START_ROW + index

            # Ambiente
            self.ws[f"B{row}"] = ambiente.nome

            # Dimensões
            self.ws[f"E{row}"] = ambiente.dimensoes.comprimento
            self.ws[f"F{row}"] = ambiente.dimensoes.largura
            self.ws[f"G{row}"] = ambiente.dimensoes.altura
            self.ws[f"H{row}"] = ambiente.dimensoes.espessura

            # Vão
            self.ws[f"J{row}"] = ambiente.vao.tipo
            self.ws[f"K{row}"] = ambiente.vao.comprimento
            self.ws[f"L{row}"] = ambiente.vao.altura
            self.ws[f"M{row}"] = ambiente.vao.espessura

            # Alvenaria adicional
            self.ws[f"O{row}"] = ambiente.alvenaria_adicional.tipo
            self.ws[f"P{row}"] = ambiente.alvenaria_adicional.comprimento
            self.ws[f"Q{row}"] = ambiente.alvenaria_adicional.altura
            self.ws[f"R{row}"] = ambiente.alvenaria_adicional.espessura


# gerador de memorial
class MemorialGenerator:
    SHEET_NAME = "Levantamento Campo"

    def __init__(self, template_path: str | Path):
        self.template_path = Path(template_path)

    def generate(self, projeto: ProjetoMemorial, output_path: str | Path) -> Path:
        ProjetoValidator.validar(projeto)

        if not self.template_path.exists():
            raise FileNotFoundError(f"Template não encontrado: {self.template_path}")

        wb = load_workbook(self.template_path)

        if self.SHEET_NAME not in wb.sheetnames:
            raise MemorialValidationError(
                f"A planilha '{self.SHEET_NAME}' não foi encontrada no template."
            )

        ws = wb[self.SHEET_NAME]
        mapper = LevantamentoCampoMapper(ws)

        mapper.preencher_titulo_projeto(projeto.nome_projeto)
        mapper.preencher_ambientes(projeto.ambientes)

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return output_path


# dados de exemplo para teste
def exemplo_projeto() -> ProjetoMemorial:
    return ProjetoMemorial(
        nome_projeto="Prj Estande de Tiro (50 m) – 5º BIL (Amv)",
        ambientes=[
            Ambiente(
                nome="Sala de Instrução",
                dimensoes=Dimensoes(
                    comprimento=10.0,
                    largura=5.0,
                    altura=3.0,
                    espessura=0.15
                ),
                vao=Vao(
                    tipo="Porta",
                    comprimento=0.90,
                    altura=2.10,
                    espessura=0.15
                ),
                alvenaria_adicional=AlvenariaAdicional(
                    tipo="Painel",
                    comprimento=2.0,
                    altura=3.0,
                    espessura=0.15
                )
            ),
            Ambiente(
                nome="Sanitário",
                dimensoes=Dimensoes(
                    comprimento=3.0,
                    largura=2.0,
                    altura=2.8,
                    espessura=0.15
                ),
                vao=Vao(
                    tipo="Porta",
                    comprimento=0.80,
                    altura=2.10,
                    espessura=0.15
                ),
                alvenaria_adicional=AlvenariaAdicional(
                    tipo="Divisória",
                    comprimento=1.2,
                    altura=2.2,
                    espessura=0.10
                )
            ),
        ]
    )



if __name__ == "__main__":
    #aqui é o caminho para o template memorial exemplar que vai ser utilizado e preenchido
    TEMPLATE_PATH = "EchoCAD/app/backend/src/templates/Memorial de Cálculo - Modelo.xlsx"
    #aqui é o caminho onde os arquivos do memorial vão ficar depois de gerados
    OUTPUT_PATH = "EchoCAD/app/backend/src/templates/saida/memorial_preenchido.xlsx"
    projeto = exemplo_projeto()
    generator = MemorialGenerator(TEMPLATE_PATH)
    arquivo_gerado = generator.generate(projeto, OUTPUT_PATH)

    print(f"Memorial gerado com sucesso em: {arquivo_gerado}")