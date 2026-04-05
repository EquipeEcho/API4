import logging
import os
from pathlib import Path
from typing import List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from dxf_extractor import Ambiente, CADExtractor, ProjetoMemorial
from sinapi import buscar_preco_sinapi, carregar_sinapi

from IA_config import *

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


MAP_SINAPI = {
    "parede": "bloco cerâmico",
    "vao": "porta madeira",
    "hidro": "tubo pvc",
    "cobertura": "telha cerâmica",
    "parede_leve": "drywall",
    "pilar": "concreto estrutural",
    "viga": "concreto armado"
}


# ==============================
# GERAÇÃO EXCEL
# ==============================
class LevantamentoCampoMapper:
    START_ROW = 8
    END_ROW = 27
    PRECO_UNITARIO_COL = "T"
    PRECO_TOTAL_COL = "U"

    def __init__(self, worksheet: Worksheet):
        self.ws = worksheet

    def preencher_titulo_projeto(self, nome_projeto: str) -> None:
        self.ws["B1"] = nome_projeto
        self.ws[f"{self.PRECO_UNITARIO_COL}7"] = "Preço Unitário SINAPI"
        self.ws[f"{self.PRECO_TOTAL_COL}7"] = "Preço Total SINAPI"

    def preencher_ambientes(self, ambientes: List[Ambiente]) -> None:
        limite = self.END_ROW - self.START_ROW + 1

        if len(ambientes) > limite:
            logger.warning(
                f"Template suporta {limite} ambientes. "
                f"Serão exportados somente os {limite} primeiros de {len(ambientes)}."
            )

        for index, ambiente in enumerate(ambientes[:limite]):
            row = self.START_ROW + index

            self.ws[f"B{row}"] = ambiente.nome
            self.ws[f"E{row}"] = ambiente.dimensoes.comprimento or 0
            self.ws[f"G{row}"] = ambiente.dimensoes.altura
            self.ws[f"H{row}"] = ambiente.dimensoes.espessura

            self.ws[f"J{row}"] = ambiente.vao.tipo
            self.ws[f"K{row}"] = ambiente.vao.comprimento or 0

            self.ws[f"L{row}"] = ambiente.area_parede
            self.ws[f"M{row}"] = ambiente.area_liquida

            self.ws[f"{self.PRECO_UNITARIO_COL}{row}"] = getattr(
                ambiente, 'custo_unitario', 0)
            self.ws[f"{self.PRECO_TOTAL_COL}{row}"] = getattr(
                ambiente, 'custo_total', 0)


class MemorialGenerator:
    SHEET_NAME = "Levantamento Campo"

    def __init__(self, template_path: str):
        self.template_path = Path(template_path)

    def generate(self, projeto: ProjetoMemorial, output_path: str) -> Path:
        wb = load_workbook(self.template_path)
        ws = wb[self.SHEET_NAME]
        mapper = LevantamentoCampoMapper(ws)
        mapper.preencher_titulo_projeto(projeto.nome_projeto)
        mapper.preencher_ambientes(projeto.ambientes)

        out_p = Path(output_path)
        out_p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_p)
        return out_p

# ==============================
# EXECUÇÃO
# ==============================


def run_integration(dxf_file: str, template_file: str, output_file: str):
    logger.info(f"Processando {os.path.basename(dxf_file)}...")

    basepath = Path(__file__).parent
    sinapi_path = basepath / "sinapi.xlsx"

    sinapi = carregar_sinapi(sinapi_path)

    extractor = CADExtractor(dxf_file)
    ambientes = extractor.extrair_dados_reais()
    logger.info(f"Ambientes encontrados: {len(ambientes)}")

    for a in ambientes:
        logger.info(f"{a.nome} {a.area_liquida} {a.tipo}")

    for ambiente in ambientes:

        categoria = ambiente.tipo

        material = MAP_SINAPI.get(categoria)

        if not material:
            logger.warning(
                f"[SINAPI] Ambiente '{ambiente.nome}' sem mapeamento de material para categoria '{categoria}'."
            )
            continue

        preco = buscar_preco_sinapi(material, sinapi)

        if preco is not None:
            ambiente.custo_unitario = preco
            ambiente.custo_total = preco * (ambiente.area_liquida or 0)
            logger.info(
                f"[SINAPI] Ambiente '{ambiente.nome}' | produto '{material}' | "
                f"preço unitário {ambiente.custo_unitario:.6f} | total {ambiente.custo_total:.6f}"
            )
        else:
            ambiente.custo_unitario = 0
            ambiente.custo_total = 0
            logger.warning(
                f"[SINAPI] Produto '{material}' não encontrado para ambiente '{ambiente.nome}'."
            )

    nome_projeto = os.path.basename(dxf_file).replace(".dxf", "").title()
    projeto = ProjetoMemorial(nome_projeto=nome_projeto, ambientes=ambientes)

    generator = MemorialGenerator(template_file)
    arquivo_final = generator.generate(projeto, output_file)

    logger.info(f"Sucesso! Memorial gerado em: {arquivo_final}")

if __name__ == "__main__":
    basepath = Path(__file__).parent
    dxf_file = (basepath / "teste.dxf").resolve()
    template_file = (basepath / "model_memorial.xlsx").resolve()
    output_file = (basepath / "memorial_preenchido.xlsx").resolve()

    run_integration(
        dxf_file=str(dxf_file),
        template_file=str(template_file),
        output_file=str(output_file)
    )
